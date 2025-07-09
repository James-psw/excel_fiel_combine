import os
import pandas as pd
from openpyxl import load_workbook

# 1. 获取当前目录下所有xlsx文件（排除临时文件）
excel_files = [f for f in os.listdir('.') if f.endswith('.xlsx') and not f.startswith('~')]

# 2. 用于存放所有入库和出库数据
data_ruku = []
data_chuku = []

# 3. 遍历每个文件，处理sheet
for file in excel_files:
    print(f"正在处理文件: {file}")
    try:
        wb = load_workbook(file, read_only=True)
        #print(f"  sheet列表: {wb.sheetnames}")
        for sheet_name in ['入库', '出库']:
            if sheet_name in wb.sheetnames:
                df = pd.read_excel(file, sheet_name=sheet_name, engine='openpyxl')
                #print(f"    {sheet_name} 列名: {list(df.columns)}")
                # 只保留第一列为数字，且第二列非空的行
                first_col = df.columns[0]
                second_col = df.columns[1] if len(df.columns) > 1 else None
                if second_col:
                    df = df[pd.to_numeric(df[first_col], errors='coerce').notna() & df[second_col].notna()]
                else:
                    df = df[pd.to_numeric(df[first_col], errors='coerce').notna()]
                if sheet_name == '入库':
                    data_ruku.append(df)
                else:
                    data_chuku.append(df)
            else:
                print(f"  没有找到sheet: {sheet_name}")
    except Exception as e:
        print(f"处理文件 {file} 时出错: {e}")

# 4. 合并所有数据
if data_ruku:
    df_ruku = pd.concat(data_ruku, ignore_index=True)
else:
    df_ruku = pd.DataFrame()

if data_chuku:
    df_chuku = pd.concat(data_chuku, ignore_index=True)
else:
    df_chuku = pd.DataFrame()

# 5. 新建exported文件夹
export_dir = 'exported'
os.makedirs(export_dir, exist_ok=True)

# 6. 保存到新的Excel文件
output_path = os.path.join(export_dir, '合并结果.xlsx')
if df_ruku.empty and df_chuku.empty:
    print("没有可合并的数据，未生成Excel文件。")
else:
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        if not df_ruku.empty:
            df_ruku.to_excel(writer, sheet_name='入库', index=False)
        if not df_chuku.empty:
            df_chuku.to_excel(writer, sheet_name='出库', index=False)
    print(f"合并完成，文件已保存到: {output_path}")
